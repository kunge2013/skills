# [AGC:FILE] tool=Cc author=fangkun date=2026-06-30
"""
读取 Excel 指定行，查询 MySQL，组装完整提示词并输出。
用法:
  python query_data.py --file <excel路径> --row <行号，从0开始>
  python query_data.py --prod-inst-id <ID> --acct-item-type-id <TYPE_ID>
"""
import os
import sys
import argparse
from datetime import datetime, date
from decimal import Decimal

import pymysql
import openpyxl
from dotenv import load_dotenv

# 修复 Windows 控制台编码问题
sys.stdout.reconfigure(encoding='utf-8', errors='replace')


# [AGC:START] tool=Cc author=fangkun
def load_db_config():
    """从 .env 文件加载数据库配置"""
    env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), '.env')
    if not os.path.exists(env_path):
        print(f"错误: .env 文件不存在: {env_path}", file=sys.stderr)
        sys.exit(1)
    load_dotenv(env_path)
    return {
        'host': os.getenv('DB_HOST', 'localhost'),
        'port': int(os.getenv('DB_PORT', '3306')),
        'user': os.getenv('DB_USER', 'root'),
        'password': os.getenv('DB_PASSWORD', ''),
        'database': os.getenv('DB_NAME', ''),
        'charset': 'utf8mb4',
    }
# [AGC:END]


# [AGC:START] tool=Cc author=fangkun
def read_excel_row(file_path, row_index):
    """读取 Excel 指定行，返回 (prod_inst_id, acct_item_type_id)"""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))  # 跳过表头
    wb.close()
    if row_index >= len(rows):
        print(f"错误: 行号 {row_index} 超出范围，共 {len(rows)} 行数据", file=sys.stderr)
        sys.exit(1)
    row = rows[row_index]
    return str(row[0]).strip(), str(row[1]).strip()
# [AGC:END]


# [AGC:START] tool=Cc author=fangkun
def get_excel_total_rows(file_path):
    """获取 Excel 数据总行数（不含表头）"""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    return len(rows)
# [AGC:END]


# [AGC:START] tool=Cc author=fangkun
def query_db(prod_inst_id, acct_item_type_id):
    """查询数据库，返回 (records, logs)"""
    config = load_db_config()
    connection = pymysql.connect(
        host=config['host'],
        port=config['port'],
        user=config['user'],
        password=config['password'],
        database=config['database'],
        charset=config['charset'],
        cursorclass=pymysql.cursors.DictCursor
    )
    try:
        with connection.cursor() as cursor:
            sql_record = """
                SELECT a.ACCT_ITEM_TYPE_ID, a.ID, a.PROD_INST_ID, b.NAME,
                       a.START_DATE, a.END_DATE, a.START_FLAG, a.LATEST_FLAG,
                       a.LOOP_MONEY, a.CAL_ACCT_RECORD_ID, a.ACCT_ID,
                       a.CREATE_DATE, a.UPDATE_DATE
                FROM cal_acct_record a
                LEFT JOIN acct_item_type b ON a.ACCT_ITEM_TYPE_ID = b.ACCT_ITEM_TYPE_ID
                WHERE a.PROD_INST_ID = %s AND a.ACCT_ITEM_TYPE_ID = %s
                ORDER BY a.START_DATE ASC
            """
            cursor.execute(sql_record, (prod_inst_id, acct_item_type_id))
            records = cursor.fetchall()

            sql_log = """
                SELECT PROD_INST_ID, BEGIN_DATE, INPUT_DATE, ATTR_ID, ATTR_NAME,
                       MOD_BEFORE, MOD_AFTER, MOD_BEFORE_VAL, MOD_AFTER_VAL,
                       MOD_DATE, MOD_REASON
                FROM prod_inst_log
                WHERE prod_inst_id = %s
                ORDER BY BEGIN_DATE DESC
            """
            cursor.execute(sql_log, (prod_inst_id,))
            logs = cursor.fetchall()
        return records, logs
    finally:
        connection.close()
# [AGC:END]


# [AGC:START] tool=Cc author=fangkun
def format_value(v):
    """格式化单个值"""
    if v is None:
        return "None"
    if isinstance(v, (datetime, date)):
        return f"'{v.strftime('%Y-%m-%d %H:%M:%S')}'"
    if isinstance(v, Decimal):
        return str(int(v)) if v == int(v) else str(v)
    if isinstance(v, (int, float)):
        return str(v)
    return f"'{v}'"


def format_data_for_prompt(records):
    """复刻 _format_data_for_prompt: key=value 格式"""
    if not records:
        return "无数据"
    lines = []
    for i, row in enumerate(records, 1):
        row_data = []
        for k in ['ID', 'PROD_INST_ID', 'ACCT_ITEM_TYPE_ID', 'NAME',
                  'START_DATE', 'END_DATE', 'START_FLAG', 'LATEST_FLAG']:
            if k in row:
                row_data.append(f"{k}={format_value(row[k])}")
        lines.append(f"  行{i}: " + ", ".join(row_data))
    return "\n".join(lines)


def format_log_data_for_prompt(logs):
    """复刻 _format_log_data_for_prompt: markdown 表格格式"""
    if not logs:
        return "无日志数据"
    headers = ['BEGIN_DATE', 'ATTR_ID', 'ATTR_NAME', 'MOD_BEFORE', 'MOD_AFTER',
               'MOD_BEFORE_VAL', 'MOD_AFTER_VAL', 'MOD_REASON']
    lines = []
    lines.append("| " + " | ".join(headers) + " |")
    separator = " | ".join(["-------------------"] * len(headers))
    lines.append("| " + separator + " |")
    for row in logs:
        values = []
        for h in headers:
            val = row.get(h, '')
            if val is None:
                val = '-'
            elif isinstance(val, (datetime, date)):
                val = val.strftime('%Y-%m-%d %H:%M:%S')
            elif isinstance(val, Decimal):
                val = str(val)
            else:
                val = str(val)
            values.append(val)
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)
# [AGC:END]


# [AGC:START] tool=Cc author=fangkun
def _strip_section_separator(text):
    """清理段落末尾的 --- 分隔符，避免与代码中添加的分隔符重复"""
    text = text.strip()
    if text.endswith('---'):
        text = text[:-3].strip()
    return text


def load_prompt_template():
    """从 references/prompt_template.md 加载并解析提示词模板。

    按 ## 标题切分 markdown，返回 {section_name: content} 字典。
    """
    skill_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    template_path = os.path.join(skill_root, 'references', 'prompt_template.md')

    if not os.path.exists(template_path):
        print(f"错误: 模板文件不存在: {template_path}", file=sys.stderr)
        sys.exit(1)

    try:
        with open(template_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except UnicodeDecodeError as e:
        print(f"错误: 模板文件编码解析失败: {template_path}\n  {e}", file=sys.stderr)
        sys.exit(1)

    # 按 ## 标题切分，清理每个段落末尾的 --- 分隔符
    sections = {}
    current_section = None
    current_content = []

    for line in content.split('\n'):
        if line.startswith('## '):
            if current_section:
                sections[current_section] = _strip_section_separator('\n'.join(current_content).strip())
            current_section = line[3:].strip()
            current_content = []
        else:
            current_content.append(line)

    if current_section:
        sections[current_section] = _strip_section_separator('\n'.join(current_content).strip())

    # 检查必需段落
    required_sections = ['角色设定', '表结构', 'Few-shot 示例', '输出格式要求']
    for section in required_sections:
        if section not in sections:
            print(f"警告: 模板文件缺少必需段落: '## {section}'", file=sys.stderr)
            sections[section] = ''

    return sections
# [AGC:END]


# [AGC:START] tool=Cc author=fangkun
def build_prompt(prod_inst_id, acct_item_type_id, records, logs, template_sections):
    """组装完整提示词 - 从 prompt_template.md 动态加载"""
    data_str = format_data_for_prompt(records)
    log_str = format_log_data_for_prompt(logs)

    role_intro = template_sections.get('角色设定', '')

    prompt = f"""{role_intro}

# 表结构
{template_sections.get('表结构', '')}

# 业务数据查询模板
{template_sections.get('业务数据查询模板', '')}

# 日志数据查询模板
{template_sections.get('日志数据查询模板', '')}

# Few-shot 示例
{template_sections.get('Few-shot 示例', '')}

{template_sections.get('推理规则总结', '')}

---

# 当前实例数据

## 当前异常数据 (cal_acct_record):
{data_str}

## 变更日志 (prod_inst_log):
{log_str}

{template_sections.get('输出格式要求', '')}

# 数据格式化方式

{template_sections.get('数据格式化方式', '')}"""
    return prompt
# [AGC:END]


# [AGC:START] tool=Cc author=fangkun
def main():
    parser = argparse.ArgumentParser(description='查询数据并生成完整提示词（从 prompt_template.md 动态加载）')
    parser.add_argument('--file', default=None, help='输入 Excel 文件路径')
    parser.add_argument('--row', type=int, default=None, help='Excel 行号（从0开始，不含表头）')
    parser.add_argument('--total', action='store_true', help='仅输出 Excel 总行数')
    parser.add_argument('--prod-inst-id', default=None, help='直接指定产品实例ID')
    parser.add_argument('--acct-item-type-id', default=None, help='直接指定账目项类型ID')
    args = parser.parse_args()

    # 仅查询总行数
    if args.total and args.file:
        print(get_excel_total_rows(args.file))
        return

    # 确定 prod_inst_id 和 acct_item_type_id
    if args.file is not None and args.row is not None:
        prod_inst_id, acct_item_type_id = read_excel_row(args.file, args.row)
    elif args.prod_inst_id and args.acct_item_type_id:
        prod_inst_id = args.prod_inst_id
        acct_item_type_id = args.acct_item_type_id
    else:
        print("错误: 请提供 --file + --row 或 --prod-inst-id + --acct-item-type-id",
              file=sys.stderr)
        sys.exit(1)

    # 加载提示词模板
    template_sections = load_prompt_template()

    # 查询数据库
    try:
        records, logs = query_db(prod_inst_id, acct_item_type_id)
    except pymysql.Error as e:
        print(f"数据库错误: {e}", file=sys.stderr)
        sys.exit(1)

    # 组装并输出提示词
    prompt = build_prompt(prod_inst_id, acct_item_type_id, records, logs, template_sections)
    print(prompt)
# [AGC:END]


if __name__ == '__main__':
    main()
